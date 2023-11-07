
import pandas as pd
df.read_excel("supermarket.xlsx")
print(df)
df= df[['Gender','Product Line', 'Total']]
pivot_table= df.pivot_table(index="Gender", columns= 'Product Line', values=" Total", aggfunc= 'sum' )


print(pivot_table)


pivot_table.to_excel("pivot_table.xlsx", "report", startrow=4)



#creating viz using openpyxl
#pip install openpyxl

from openpyxl import load_workbook
from openpyxl import BarChart, Reference
wb=load_workbook('pivot_table.xlsx')


sheet= wb['report']
# select active rows and columns
min_column= wb.active.min_column
max_column= wb.active.max_column
min_row= wb.active.min_row
max_row= wb.active.max_row

barchart= BarChart()

data= Reference(sheet, min_col= min_column, max_col= max_column, min_row=min_row,max_row =max_row)

categories= Reference(sheet, min_col= min_column, max_col= min_column, min_row=min_row + 1,max_row =max_row)

#creating the barchart

barchat.add_data(data, titles_from_data=True)
bachart.set_categories(categories)

sheet.add_chart(barchart, "B12")
barchart.title= "Sales by Product Line"
barchart.style= 5
wb.save('barchart.xlsx')

